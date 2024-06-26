import atexit
import zipfile
import os
from dash import Dash, dcc, html, Input, Output, State
from dash.exceptions import PreventUpdate
import dash_bootstrap_components as dbc
import pandas as pd
from dash_extensions import EventListener
from dash_bootstrap_components import Tooltip

from openpyxl import load_workbook

app = Dash(__name__, external_stylesheets=[dbc.themes.CYBORG])  # SLATE also good

# Custom CSS styles to reduce the size of H4 elements
custom_styles = {
    'h4': {
        'font-size': '1.5rem',  # Adjust the font size as needed
    }
}

# Define more custom CSS styles
textarea_style = {
    'resize': 'none',
    'width': '100%',
    'height': '150px',
    'overflow': 'auto',
    'background-color': '#373a3c',  # Match the background color of the card
    'border-color': '#495057',  # Match the border color of the card
    'color': '#fff',  # Match the text color of the card
}

# Define sheet names globally
# sheet_names = ['Inter', 'Polref', 'Offspec', 'Surf']

tab1_content = dbc.Container([
    # html.H2(children='FAP 9', style={'textAlign': 'center', 'fontSize': '1.5rem'}),  # fontsize adjusted
    dbc.Row(html.Br()),
    # Row 1
    html.Div(
        [
            dbc.Row(
                [
                    dbc.Col(dcc.Dropdown(
                        options=[],  # [{"label": i, "value": i} for i in ['All'] + sheet_names],
                        id="instrument", value='All', style={'height': '30px'}),
                        width=2),
                    dbc.Col(html.H4(children='', id='title-link', style={
                        'textAlign': 'left',
                        'fontSize': '1.5rem',  # Adjust the font size as needed
                        'whiteSpace': 'nowrap',
                        'overflow': 'hidden',
                        'textOverflow': 'ellipsis',
                        'maxWidth': '100%',
                        'width': '100%'  # Ensure the element takes the full width of the container
                    }, title=''), width=10)
                ]
            ),
            dbc.Row(
                [
                    dbc.Col(dcc.Dropdown(
                        id="rb-numbers", style={'height': '30px'}), width=2, style={'font-size': '16px'}),
                    dbc.Col(html.Label("PI: "), width=1, style={'textAlign': 'right'}),
                    dbc.Col(dcc.Input(id="pi-name", type="text", placeholder="Prof. U.N. Known", readOnly=True,
                                      debounce=True, style={'height': '30px'}), width=3),
                    dbc.Col(html.Div("Days: "), width=1, style={'textAlign': 'right'}),
                    dbc.Col(dcc.Input(id="days", type="text", placeholder="", readOnly=True, debounce=True,
                                      style={'width': '60px', 'height': '30px'}), width=2),
                    # Adding Tooltip
                    dbc.Tooltip(
                        "Use Ctrl+Left Arrow and Ctrl+Right Arrow to navigate.",
                        target="rb-numbers",
                        placement="top",
                        delay={'show': 0, 'hide': 20}
                    ),
                ],
                justify="start"
            ),
            dbc.Row(
                [
                    dbc.Col(html.Div("Speaker 1: "), width={"size": "1", "offset": 2}, style={'textAlign': 'right'}),
                    dbc.Col(
                        dcc.Input(id="speaker-1", type="text", placeholder="Speaker One", readOnly=True, debounce=True,
                                  style={'height': '30px'}), width=3),
                    dbc.Col(html.Div("Score 1: "), width=1, style={'textAlign': 'right'}),
                    dbc.Col(dcc.Input(id="score-1", type="text", placeholder="7.5", readOnly=True, debounce=True,
                                      style={'width': '60px', 'height': '30px'}), width=2),
                ],
                justify="start"
            ),
            dbc.Row(
                [
                    dbc.Col(html.Div("Speaker 2: "), width={"size": "1", "offset": 2}, style={'textAlign': 'right'}),
                    dbc.Col(
                        dcc.Input(id="speaker-2", type="text", placeholder="Speaker Two", readOnly=True, debounce=True,
                                  style={'height': '30px'}), width=3),
                    dbc.Col(html.Div("Score 2: "), width=1, style={'textAlign': 'right'}),
                    dbc.Col(dcc.Input(id="score-2", type="text", placeholder="5.5", readOnly=True, debounce=True,
                                      style={'width': '60px', 'height': '30px'}), width=2),
                    dbc.Col(html.Div("Final score: "), width=1, style={'textAlign': 'right'}),
                    dbc.Col(dcc.Input(id="final-score", type="text", placeholder="6.5", debounce=True,
                                      style={'width': '60px', 'height': '30px'}), width=2),
                ],
                justify="start"
            ),
            # dbc.Row(html.Br()),
            dbc.Row(
                [
                    dbc.Col(html.Div("Comment 1: ")),
                    dbc.Col(dbc.Card(dbc.CardBody(id="comment-1", style={'height': '150px', 'overflow': 'auto'})),
                            width=12)
                ],
                justify="start"
            ),
            dbc.Row(
                [
                    dbc.Col(html.Div("Comment 2: ")),
                    dbc.Col(dbc.Card(dbc.CardBody(id="comment-2", style={'height': '150px', 'overflow': 'auto'})),
                            width=12)
                ],
                justify="start"
            ),

            dbc.Row(html.Br()),
            dbc.Row(
                [
                    dbc.Col(html.Div("Technical comments: "), align='right'),
                    dbc.Col(dbc.Card(dbc.CardBody(id="tech-comments", style={'height': '150px', 'overflow': 'auto'})),
                            width=5),
                    dbc.Col(html.Div("Feedback comments: ")),
                    dbc.Col(dbc.Textarea(id="feedback", style=textarea_style),
                            # {'height': '150px', 'width': '100%', 'overflow': 'auto'}
                            width=5)
                ],
                justify="start"
            ),
            dbc.Row(html.Br()),
            dbc.Row(
                [
                    dbc.Col(html.Div(id='confirmation', style={'color': 'green'})),
                    dbc.Col(html.Div(id='confirmation2', style={'color': 'green'}))
                ]
            )
        ]
    ),
])

tab2_content = dbc.Card([
    html.Div("Scoring portal file: "),
    dbc.Input(id='portal-file', type="text", placeholder="Downloaded Excel file from scoring portal",
              debounce=True, style={'height': '30px'}),

    html.Div("Final scores file: "),
    dbc.Input(id='scores-file', type="text", placeholder="Scores file",
              debounce=True, style={'height': '30px'})
]
)

tabs = dbc.Tabs(
    [
        dbc.Tab(tab1_content, label="FAP 9 Dashboard", activeTabClassName="fw-bold fst-italic"),
        dbc.Tab(tab2_content, label="File settings", activeTabClassName="fw-bold fst-italic"),
    ]
)

# app.layout = tabs

app.layout = html.Div([
    tabs,
    EventListener(
        id="event-listener",
        events=[{"event": "keydown", "props": ["ctrlKey", "key"]}]
    )
])


@app.callback(
    Output('rb-numbers', 'value'),
    Input('event-listener', 'n_events'),
    State('rb-numbers', 'options'),
    State('rb-numbers', 'value'),
    State('event-listener', 'event')
)
def handle_key_event(n_events, options, current_value, event):
    if n_events is None or not event or event['ctrlKey'] is False:
        raise PreventUpdate

    if not options or not current_value:
        raise PreventUpdate

    # Find the index of the current value
    current_index = next((index for (index, d) in enumerate(options) if d["value"] == current_value), None)

    if current_index is None:
        raise PreventUpdate

    # Move to the next value if ArrowRight is pressed
    if event['key'] == 'ArrowRight':
        next_index = (current_index + 1) % len(options)
    # Move to the previous value if ArrowLeft is pressed
    elif event['key'] == 'ArrowLeft':
        next_index = (current_index - 1) % len(options)
    else:
        raise PreventUpdate

    return options[next_index]['value']


@app.callback(
    Output('instrument', 'options'),
    Input('portal-file', 'value')
)
def fill_dropdown(portal_file):
    if not portal_file:
        raise PreventUpdate

    xls = pd.ExcelFile(portal_file)
    sheetnames = xls.sheet_names
    print(sheetnames.sort())
    return sheetnames


@app.callback(
    Output('rb-numbers', 'options'),
    Input('instrument', 'value'),
    State('portal-file', 'value'),
    State('scores-file', 'value')

)
def update_rb_numbers_options(selected_instrument, portal_file, scores_file):
    if not portal_file or not scores_file:
        raise PreventUpdate

    # Load the new data
    df_list = []
    xls = pd.ExcelFile(portal_file)

    for sheet_name in xls.sheet_names:
        df = pd.read_excel(portal_file, sheet_name=sheet_name)
        df['Instrument'] = sheet_name
        df_list.append(df)
    combined_df = pd.concat(df_list, ignore_index=True)

    if selected_instrument == 'All':
        rb_options = [{'label': str(rb), 'value': str(rb)} for rb in combined_df['Proposal Reference Number'].unique()]
    else:
        rb_options = [{'label': str(rb), 'value': str(rb)} for rb in
                      combined_df[combined_df['Instrument'] == selected_instrument][
                          'Proposal Reference Number'].unique()]
    sorted_rb_options = sorted(rb_options, key=lambda x: x["label"])
    return sorted_rb_options


@app.callback(
    Output('title-link', 'children'),
    Output('title-link', 'title'),
    Output('pi-name', 'value'),
    Output('days', 'value'),
    Output('speaker-1', 'value'),
    Output('score-1', 'value'),
    Output('speaker-2', 'value'),
    Output('score-2', 'value'),
    Output('final-score', 'value'),
    Output('comment-1', 'children'),
    Output('comment-2', 'children'),
    Output('instrument', 'value'),
    Output('feedback', 'value'),
    Input('rb-numbers', 'value'),
    State('instrument', 'value'),
    State('portal-file', 'value'),
    State('scores-file', 'value')
)
def update_details(rb_number, selected_instrument, portal_file, scores_file):
    if rb_number is None or not portal_file or not scores_file:
        raise PreventUpdate

    # Load the new data
    df_list = []
    xls = pd.ExcelFile(portal_file)
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(portal_file, sheet_name=sheet_name)
        df['Instrument'] = sheet_name
        df_list.append(df)
    combined_df = pd.concat(df_list, ignore_index=True)

    # Filter the combined DataFrame for the selected Proposal Reference Number
    df_filtered = combined_df[combined_df['Proposal Reference Number'] == int(rb_number)]

    if df_filtered.empty:
        return '', '', '', '', '', '', '', 'N/A', 'No comments available', 'No comments available', selected_instrument

    title = df_filtered['Proposal Title'].values[0] if 'Proposal Title' in df_filtered.columns else 'N/A'
    pi_name = df_filtered['Principal Investigator'].values[
        0] if 'Principal Investigator' in df_filtered.columns else 'N/A'
    days = df_filtered['Requested Time'].values[0] if 'Requested Time' in df_filtered.columns else 'N/A'
    speaker_1 = df_filtered['Speaker 1'].values[0] if 'Speaker 1' in df_filtered.columns else 'N/A'
    score_1 = df_filtered['Reviewer 1 score'].values[0] if 'Reviewer 1 score' in df_filtered.columns else 'N/A'
    speaker_2 = df_filtered['Speaker 2'].values[0] if 'Speaker 2' in df_filtered.columns else 'N/A'
    score_2 = df_filtered['Reviewer 2 score'].values[0] if 'Reviewer 2 score' in df_filtered.columns else 'N/A'
    comment_1 = df_filtered['Reviewer 1 review comment'].values[
        0] if 'Reviewer 1 review comment' in df_filtered.columns else 'No comments available'
    comment_2 = df_filtered['Reviewer 2 review comment'].values[
        0] if 'Reviewer 2 review comment' in df_filtered.columns else 'No comments available'
    instrument = df_filtered['Instrument'].values[0] if 'Instrument' in df_filtered.columns else selected_instrument

    # Extract the final score and feedback from the Excel file
    wb = load_workbook(filename=scores_file)
    sheet_ranges = wb['Proposals']
    final_score = 'N/A'
    feedback_comment = 'N/A'
    for row in sheet_ranges.iter_rows(min_row=2, max_row=sheet_ranges.max_row, values_only=True):
        if row[0] == int(rb_number):
            if len(row) > 8:
                final_score = row[8] if row[8] is not None else 'N/A'
            if len(row) > 10:
                feedback_comment = row[10] if row[10] is not None else 'N/A'
            break
    wb.close()

    # Generate the link with the Proposal Reference Number
    link = (f"https://stfc365-my.sharepoint.com/personal/emma_gozzard_stfc_ac_uk/Documents/ISIS%20FAPs/24-2/FAP-Secs"
            f"/Proposal-PDFs/{rb_number}.pdf")

    return (html.A(title, href=link, target="_blank"), title, pi_name, days, speaker_1, score_1, speaker_2, score_2,
            final_score, comment_1, comment_2, instrument, feedback_comment)


@app.callback(
    Output('confirmation', 'children'),
    Input('final-score', 'value'),
    State('rb-numbers', 'value'),
    State('scores-file', 'value')
)
def update_final_score(final_score, rb_number, scores_file):
    if not rb_number or not final_score or not scores_file:
        raise PreventUpdate

    # Update the final score in the Excel file
    wb = load_workbook(filename=scores_file)
    sheet_ranges = wb['Proposals']
    updated = False
    for row in sheet_ranges.iter_rows(min_row=2, max_row=sheet_ranges.max_row):
        if row[0].value == int(rb_number):
            row[8].value = final_score
            updated = True
            break

    if updated:
        wb.save(scores_file)
        wb.close()
        return "Final score updated successfully!"
    else:
        wb.close()
        return "Error: Proposal Reference Number not found."


@app.callback(
    Output('confirmation2', 'children'),
    Input('feedback', 'value'),
    State('rb-numbers', 'value'),
    State('scores-file', 'value')
)
def update_feedback(feedback, rb_number, scores_file):
    if not rb_number or not feedback or not scores_file:
        raise PreventUpdate

    if not os.path.exists(scores_file):
        return "Error: Scores file does not exist."

    try:
        # Update the feedback in the Excel file
        wb = load_workbook(filename=scores_file)
        sheet_ranges = wb['Proposals']
        updated = False
        for row in sheet_ranges.iter_rows(min_row=2, max_row=sheet_ranges.max_row):
            if row[0].value == int(rb_number):
                row[10].value = feedback
                updated = True
                break

        if updated:
            wb.save(scores_file)
            wb.close()
            return "Feedback updated successfully!"
        else:
            wb.close()
            return "Error: Proposal Reference Number not found."
    except zipfile.BadZipFile:
        return "Error: The scores file is corrupted or incomplete."
    except Exception as e:
        return f"Error: {str(e)}"


if __name__ == '__main__':
    app.run_server(debug=True)
